import openpyxl
from datetime import datetime


def write_lines(input_file, header):
    booking_lines = header
    if input_file.endsWith == '.xlsx':
        try:
            workbook = openpyxl.load_workbook(input_file)
            sheet = workbook.active
            max_rows = sheet.max_row
            for r in range(2, max_rows + 1):
                try:
                    valuate_date = datetime.strptime(sheet.cell(row=r, column=2).value, '%m/%d/%Y').strftime('%Y-%m-%dT%H:%M')
                    note = sheet.cell(row=r, column=3).value
                    amount = str(sheet.cell(row=r, column=7).value).replace('.', ',')
                    if sheet.cell(row=r, column=5).value is not None:
                        note = note + sheet.cell(row=r, column=5).value
                    booking_type = ''
                    if note.upper().startswith('BETRAG DER EINGEZ'):
                        booking_type = 'Einlage'
                    elif note.upper().startswith('ZAHLUNG FÜR DEN ZINS'):
                        booking_type = 'Zinsen'
                    if booking_type != '':
                        booking_lines.append(valuate_date + ';' + booking_type + ';' + amount + ';EUR;' + note + '\r\n')
                except Exception as inner_error:
                    print('Fehler bei der Verarbeitung der Zeile ' + str(r) + ': ' + repr(inner_error))
        except Exception as error:
            print('Allgemeiner Fehler beim Vearbeiten der XLSL Datei: ' + repr(error))
    else:
        print('Die angegeben Dateieindung wird nicht unterstütz.')
        raise ValueError('Die Dateieindung wird nicht unterstütz.')
    return booking_lines
