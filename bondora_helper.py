# coding=utf-8
import openpyxl


def write_lines(input_file, header):
    booking_lines = header
    if input_file.lower().endswith('.xlsx'):
        try:
            workbook = openpyxl.load_workbook(input_file)
            sheet = workbook.active
            max_rows = sheet.max_row
            for r in range(2, max_rows + 1):
                try:
                    valuate_date = sheet.cell(row=r, column=1).value.strftime('%Y-%m-%dT%H:%M')
                    if sheet.cell(row=r, column=6).value is not None:
                        note = 'Darlehen ' + sheet.cell(row=r, column=6).value
                    else:
                        note = 'Einzahlung'
                    amount = str(sheet.cell(row=r, column=3).value).replace('.', ',')
                    description = str(sheet.cell(row=r, column=5).value).replace('.', ',')
                    booking_type = ''
                    if description.lower().startswith('transferdeposit|'):
                        booking_type = 'Einlage'
                    elif description.lower().startswith('transferinterestrepaiment'):
                        booking_type = 'Zinsen'
                    elif description.lower().startswith('transfergogrow'):
                        booking_type = 'Umbuchung'
                    if booking_type != '':
                        booking_lines.append(valuate_date + ';' + booking_type + ';' + amount + ';EUR;' + note + '\r\n')
                except Exception as inner_error:
                    print('Fehler bei der Verarbeitung der Zeile ' + str(r) + ': ' + repr(inner_error))
        except Exception as error:
            print('Allgemeiner Fehler beim Verarbeiten der XLSX Datei: ' + repr(error))
    else:
        print('Die angegeben Dateiendung wird nicht unterstützt.')
        raise ValueError('Die Dateiendung wird nicht unterstützt.')
    return booking_lines
