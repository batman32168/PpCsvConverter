import column
import uuid
import xlrd
import openpyxl
from collections import OrderedDict
from datetime import datetime, timedelta

from column_types import column_types
from summary_types import summary_types


class extractor:
    def __init__(self, configuration: str):
        self._configuration = configuration

    @property
    def splitt_char(self):
        if self._configuration['splitt_char'] is None:
            return ';'
        else:
            return self._configuration['splitt_char']

    @property
    def start_row(self):
        if self._configuration['start_row'] is None:
            return 0
        else:
            return self._configuration['start_row']

    @property
    def filepattern(self):
        if self._configuration['filepattern'] is None:
            return '*.*'
        else:
            return self._configuration['filepattern']

    @property
    def name(self):
        if self._configuration['name'] is None:
            return uuid.UUID
        else:
            return self._configuration['name']

    @property
    def summary(self):
        if 'summary' not in self._configuration or self._configuration['summary'] is None or self._configuration['summary'] =='' :
            return summary_types.single
        else:
            temp = self._configuration['summary'].lower()
            if temp == 'daily':
                return summary_types.daily
            if temp == 'weekly':
                return summary_types.weekly
            if temp == 'monthly':
                return summary_types.daily
            return summary_types.single

    @property
    def columns(self):
        temp_columns: OrderedDict ={}
        if self._configuration['columns'] is not None:
            for column_config in self._configuration['columns']:
                column_name = column_config['name']
                temp_class =column.column(column_config)
                temp_columns[column_name] = temp_class
        return temp_columns

    def extract_lines(self, file: str):
        transactions = []
        if self.filepattern.lower().endswith('.xls'):
            transactions = self._extract_xls_file(file)
        elif self.filepattern.lower().endswith('.xlsx'):
            transactions = self._extract_xlsx_file(file)
        elif self.filepattern.lower().endswith('.csv'):
            transactions = self._extract_csv_file(file)

        return self._convert_transactions(transactions)

    def _extract_xls_file(self, input_file: str):
        transaction = []
        if input_file.lower().endswith('.xls'):
            try:
                workbook = xlrd.open_workbook(input_file)
                print('Datei erfolgreich geöffnet')
                sheet = workbook.sheet_by_index(0)
                max_rows = sheet.nrows
                columns = self.columns.values()
                for r in range(self.start_row-1, max_rows):
                    print('Verarbeite Zeile ' + str(r))
                    try:
                        line = {}
                        for column in columns:
                            cell_value = column.get_formated_value(sheet.cell_value(rowx=r, colx=column.column_number-1))
                            line[column.header.name] = cell_value
                        transaction.append(line)
                    except ValueError:
                        print('Keine Buchungstyp gefunden --> Zeile wird nicht importiert')
                    except Exception as inner_error:
                        print('Fehler bei der Verarbeitung der Zeile ' + str(r) + ': ' + repr(inner_error))
            except Exception as error:
                print('Allgemeiner Fehler beim Verarbeiten der XLS Datei: ' + repr(error))
        return transaction

    def _extract_xlsx_file(self, input_file: str):
        transaction = []
        if input_file.lower().endswith('.xlsx'):
            try:
                workbook = openpyxl.load_workbook(input_file)
                print('Datei erfolgreich geöffnet')
                sheet = workbook.active
                max_rows = sheet.max_row
                columns = self.columns.values()
                for r in range(self.start_row, max_rows):
                    print('Verarbeite Zeile ' + str(r))
                    try:
                        line ={}
                        for column in columns:
                            cell_value = column.get_formated_value(sheet.cell(row=r, column=column.column_number).value)
                            line[column.header.name] = cell_value
                        transaction.append(line)
                    except ValueError:
                        print('Keine Buchungstyp gefunden --> Zeile wird nicht importiert')
                    except Exception as inner_error:
                        print('Fehler bei der Verarbeitung der Zeile ' + str(r) + ': ' + repr(inner_error))
            except Exception as error:
                print('Allgemeiner Fehler beim Verarbeiten der XLSX Datei: ' + repr(error))
        return transaction

    def _convert_transactions(self, transactions):
        result =[]
        temp_helper = {}
        for tran in transactions:
            line_date = tran[column_types.valuta_date.name]
            if self.summary == summary_types.daily:
                line_date = datetime.strptime(tran[column_types.valuta_date.name], '%Y-%m-%d %H:%M').strftime('%Y-%m-%d')
            elif self.summary == summary_types.weekly:
                line_date = self._last_day_of_week(datetime.strptime(tran[column_types.valuta_date.name], '%Y-%m-%d %H:%M')).strftime('%Y-%m-%d')
            elif self.summary == summary_types.monthly:
                line_date = self._last_day_of_month(datetime.strptime(tran[column_types.valuta_date.name], '%Y-%m-%d %H:%M')).strftime('%Y-%m-%d')

            search_key= str.format("{}-{}",line_date,tran[column_types.booking_type.name])
            if search_key in temp_helper:
                temp_helper[search_key] = self._update_transaction(temp_helper[search_key],tran)
            else:
                tran[column_types.valuta_date.name] = line_date
                temp_helper[search_key] = tran
        for summary in temp_helper.values():
            line=''
            for key in summary.keys():
                if summary[key] is not None:
                    line = line+str(summary[key])+";"
                else:
                    line = line + ";;"
            line +='\r'
            result.append(line)
        return result

    def _update_transaction(self, current, add):
        for key in current.keys():
            if key != 'valuta_date':
                try:
                    if add[key] is not None:
                        current[key] = float(current[key])+float(add[key])
                except ValueError:
                    if(current[key] != add[key]):
                        current[key] = str.format('{} | {}',current[key],add[key])
        return current

    def _last_day_of_week(self, any_day):
        week_start = any_day - timedelta(days=any_day.weekday())
        week_end = week_start + timedelta(days=6)
        last_day_of_month = self._last_day_of_month(any_day)
        if week_end <= last_day_of_month:
            return week_end
        else:
            return last_day_of_month

    def _last_day_of_month(self, any_day):
        next_month = any_day.replace(day=28) + timedelta(days=4)
        return next_month - timedelta(days=next_month.day)